import os
import openpyxl
from typing import List

XLS_FILE = 'report.xlsx'
LOG_FILE = 'log.txt'


def rename_files(surname_l: List):

    os.chdir(os.path.join(os.getcwd(), 'origin'))
    files = os.listdir()

    for file in files:
        for n, surname in enumerate(surname_l):

            written_surnames = []

            if len(surname) > 3:

                # Stampare lista che faccia fede dell'associazione fra numero e cognome

                if surname not in written_surnames:
                    written_surnames.append(surname)
                    with open(LOG_FILE, 'a') as f:
                        f.write('{}     {}\n'.format(n, surname))

                # n numero progressivo che corrisponde all'ordine in cui vengono passati i cognomi
                # il controllo viene fatto sui nomi in maiuscolo
                if surname.upper() in file.upper():
                    count = 1
                    filetype = file.split('.')[1]
                    new_name = '{}_{}'.format(str(n), count)

                    # Controlla che il file rinominato non esista già nella directory

                    while os.path.isfile('./{}.{}'.format(new_name, filetype)):
                        count = count + 1
                        temp = new_name.split('_')[0]
                        new_name = '{}_{}'.format(temp, count)

                    nn_final = '{}.{}'.format(new_name, filetype)
                    # Test version
                    print(file, ' ----> ', nn_final)
                    os.rename(file, nn_final)


def read_excel(filepath: str) -> List:

    wb_obj = openpyxl.load_workbook(filepath)

    sheet_obj = wb_obj.active

    m_row = sheet_obj.max_row

    s_list: List = []

    for i in range(1, m_row + 1):
        # Il parametro column deve corrispondere alla colonna in cui sono scritti i cognomi
        cell_obj = sheet_obj.cell(row=i, column=1)
        s_list.append(cell_obj.value)

    return s_list


if __name__ == '__main__':

    cwd = os.getcwd()

    print('\nSCRIPT PER RINOMINARE FOTO\n')
    print('Directory in cui si trovano i files: {}\n'.format(cwd))
    print('ATTENZIONE!!! CONTROLLARE CHE LA DIRECTORY SIA GIUSTA!\n')

    user_input = input('Per procedere premere "Y":\n')

    if user_input == 'Y':
        surname_list = read_excel(XLS_FILE)
        print(surname_list)
        rename_files(surname_list)

